import pdfminer.high_level
import openpyxl
from openpyxl import Workbook
import sys
import os

def readMonacoPDF(file):
    #Read the pdf
    text = pdfminer.high_level.extract_text(file)
    #text = pdfminer.high_level.extract_text(sys.argv[1])

    ##Read the pdf
    #Separate all pages
    pages = text.split('\x0c')
    #print(pages)

    #Define variables:
    #  . pageCanvas is a boolean to know if we are in an odd page or in a pair page (the canvas is not the same)
    #  . mainDict will store all the read informations
    #  . indexTag represents the entry in mainDict for a column in the pdf
    #  . nbColumn is the number of columns per page (usually 4)
    #  . readIndex (defined later) is the current index of where we are reading the file
    #  . nbRow (defined later) is the number of rows in the odd page
    #  . allPageCanvasOne is a boolean. It is True if the pdf contains anly First Canvas page type
    pageCanvas = True
    mainDict = {}
    indexTag = 1
    nbColumn = 0
    nbRow = 0
    allPageCanvasOne = False

    #We will read all pages, and separate all lines, then according to the pageCanvas we have 2 different ways to read the page
    for page, pageIndex in zip(pages, range(len(pages))):
        lines = page.split('\n')
        #print(lines)

        #For odd pages, everything start with the entry Patient ID
        #We look for all LW because we have the Length1 and Length2 before that
        #We look for all "something" Start or "something" End because this is the tag name and we have just after that the angle
        #We look for fractions because just after that we have the MU. Fractions is identical for all pages and for the first page it's just before the first X2 (different place for other pages)
        #We look for the number of rows in the page. To do so, we look for the serie 1, 2, 3, 4, ... and we take the last element
        #We look for all Y because until an empty string ('') we have all Y values
        #We look for all X1 because until an empty string ('') we have all Y values
        #For X2 it's a little bit more complicated: the first X2 could not be followed by the values. In such a case, to find the values we look for nbRows float between 2 empty string. For the other X2 it is correct
        #We stored the values for nbColumn (usually 4) columns by 1 page

        if pageCanvas and 'Patient ID:' in lines: #canvas like page 1
            readIndex = 0
            readIndex = lines.index('Patient ID:')
            if not "patientId" in mainDict:
                mainDict["patientId"] = int(lines[readIndex+2])
            if not "fractions" in mainDict:
                readIndex = lines.index('X2')
                mainDict["fractions"] = int(lines[readIndex-2])
            if pageIndex == 0:
                linesSecondPage = pages[pageIndex +1].split('\n')
                if "Width2(cm) / Label" in linesSecondPage:
                    allPageCanvasOne = True
            if nbRow == 0:
                subList = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31"]
                sll=len(subList)
                for ind in (i for i,e in enumerate(lines) if e==subList[0]):
                    if lines[ind:ind+sll]==subList:
                        nbRowIndex = ind
                        break
                findEndIndex = lines[nbRowIndex+30:].index('') + nbRowIndex+30 -1
                nbRow = int(lines[findEndIndex])
                if allPageCanvasOne:
                    mainDict["nbStrips"] = nbRow
            readIndex = lines.index('Patient ID:')
            nbColumn = 0
            while readIndex < len(lines):
                while readIndex < len(lines) and not ' Start' in lines[readIndex] and not ' End' in lines[readIndex]:
                    readIndex += 1
                if readIndex == len(lines):
                    break
                mainDict[str(indexTag+nbColumn)] = {}
                mainDict[str(indexTag+nbColumn)]["Sequence"] = int(lines[readIndex-1])
                mainDict[str(indexTag+nbColumn)]["Tag"] = lines[readIndex]
                if lines[readIndex+1] != '':
                    mainDict[str(indexTag+nbColumn)]["Angle"] = float(lines[readIndex+1])
                else:
                    mainDict[str(indexTag+nbColumn)]["Angle"] = float(lines[readIndex+2])
                nbColumn += 1
                readIndex += 2
            readIndex = lines.index('Patient ID:')
            for i in range(nbColumn):
                readIndex = lines[readIndex+4:].index('LW') + readIndex+4
                mainDict[str(indexTag+i)]["Length1"] = float(lines[readIndex-8])
                mainDict[str(indexTag+i)]["Length2"] = float(lines[readIndex-7])
                mainDict[str(indexTag+i)]["Y"] = []
                mainDict[str(indexTag+i)]["X1"] = []
                mainDict[str(indexTag+i)]["X2"] = []
            readIndex = lines.index('Patient ID:')
            i = 0
            while i<nbColumn:
                readIndex = lines[readIndex+2:].index(str(mainDict["fractions"])) + readIndex+2 +1
                try:
                    int(lines[readIndex])
                    readIndex = readIndex -1
                except:
                    mainDict[str(indexTag +i)]["MU"] = float(lines[readIndex])
                    i += 1
            if not allPageCanvasOne:
                pageCanvas = not(pageCanvas)

            readIndex = lines.index('Patient ID:')
            startX2 = 10000000
            for i in range(nbColumn):
                readIndex = lines[readIndex+1:].index('Y') + readIndex+1 +1
                while readIndex < len(lines) and lines[readIndex] != '':
                    mainDict[str(indexTag +i)]["Y"] += [float(lines[readIndex])]
                    readIndex += 1
            readIndex = lines.index('Patient ID:')
            for i in range(nbColumn):
                readIndex = lines[readIndex+2:].index('X1') + readIndex+2 +1
                while readIndex < len(lines) and lines[readIndex] != '':
                    mainDict[str(indexTag +i)]["X1"] += [float(lines[readIndex])]
                    readIndex += 1
            readIndex = lines.index('Patient ID:')
            for i in range(nbColumn):
                readIndex = lines[readIndex+2:].index('X2') + readIndex+2 +1
                foundX2 = False
                if i == 0:
                    try:
                        _ = float(lines[readIndex])
                    except:
                        tempFirstReadIndex = readIndex
                        listAllEmpytString = [j+readIndex for j,val in enumerate(lines[readIndex:]) if val==""]
                        listAllEmpytString = [readIndex] + listAllEmpytString
                        j = 1
                        while j<len(listAllEmpytString):
                            if listAllEmpytString[j] - listAllEmpytString[j-1] -1 == nbRow:
                                try:
                                    _ = int(lines[listAllEmpytString[j-1]+1])
                                except:
                                    startX2 = listAllEmpytString[j-1]+1
                                    break
                            j += 1
                        if startX2 == 10000000:
                            print("X2 was not found")
                            sys.exit()
                        readIndex = startX2
                        foundX2 = True
                while readIndex < len(lines) and lines[readIndex] != '':
                    mainDict[str(indexTag +i)]["X2"] += [float(lines[readIndex])]
                    readIndex += 1
                if i == 0 and foundX2:
                    readIndex = tempFirstReadIndex

        #For pair pages, everything start with the entry Index and the following integers to know the nbStrips
        #Y, X1 and X2 are well ordered separated by an empty string (''), so it's easy
        # We just have to read everything one entry by one entry
        # We store nbStrips to check later if it's always the same number (usually 80)

        else: #canvas like page 2
            readIndex = 0

            if not 'Index ' in lines:
                indexTag += nbColumn
                pageCanvas = not(pageCanvas)
                continue
            readIndex = lines.index('Index ')
            while readIndex < len(lines):
                try:
                    _ = int(lines[readIndex])
                    break
                except:
                    readIndex += 1
            while readIndex < len(lines) and lines[readIndex] != '':
                readIndex += 1
            if not "nbStrips" in mainDict and readIndex < len(lines) and lines[readIndex] == '':
                mainDict["nbStrips"] = int(lines[readIndex-1])
            for i in range(nbColumn):
                readIndex += 1
                while readIndex < len(lines) and lines[readIndex] != '':
                    mainDict[str(indexTag+i)]["Y"] += [float(lines[readIndex])]
                    readIndex += 1
                readIndex += 1
                while readIndex < len(lines) and lines[readIndex] != '':
                    mainDict[str(indexTag+i)]["X1"] += [float(lines[readIndex])]
                    readIndex += 1
                readIndex += 1
                while readIndex < len(lines) and lines[readIndex] != '':
                    mainDict[str(indexTag+i)]["X2"] += [float(lines[readIndex])]
                    readIndex += 1
            indexTag += nbColumn
            pageCanvas = not(pageCanvas)
    #print(mainDict)

    ##Check the values
    #Check if for all entries of mainDict, the type of the value is correct type and there is nbStrip (usually 40 or 80) values for list

    if not isinstance(mainDict["nbStrips"], int):
        print("nbStrips is not correct")
        print(mainDict["nbStrips"])
    if allPageCanvasOne:
        if mainDict["nbStrips"] != 40:
            print("nbStrips is not equal to 40")
            print(mainDict["nbStrips"])
    else:
        if mainDict["nbStrips"] != 80:
            print("nbStrips is not equal to 80")
            print(mainDict["nbStrips"])
    for i in range(len(mainDict)-3):
        if mainDict[str(i+1)]["Sequence"] == "":
            print("Sequence is not correct for " + str(i+1))
            print(mainDict[str(i+1)]["Sequence"])
        if mainDict[str(i+1)]["Tag"] == "":
            print("Tag is not correct for " + str(i+1))
            print(mainDict[str(i+1)]["Tag"])
        if len(mainDict[str(i+1)]["Y"]) != mainDict["nbStrips"]:
            print("Y is not correct for " + mainDict[str(i+1)]["Tag"])
            print(mainDict[str(i+1)]["Y"])
        if len(mainDict[str(i+1)]["X1"]) != mainDict["nbStrips"]:
            print("X1 is not correct for " + mainDict[str(i+1)]["Tag"])
            print(mainDict[str(i+1)]["X1"])
        if len(mainDict[str(i+1)]["X2"]) != mainDict["nbStrips"]:
            print("X2 is not correct for " + mainDict[str(i+1)]["Tag"])
            print(mainDict[str(i+1)]["X2"])
        if not isinstance(mainDict[str(i+1)]["MU"], float):
            print("MU is not correct for " + mainDict[str(i+1)]["Tag"])
            print(mainDict[str(i+1)]["MU"])
        if not isinstance(mainDict[str(i+1)]["Angle"], float):
            print("Angle is not correct for " + mainDict[str(i+1)]["Tag"])
            print(mainDict[str(i+1)]["Angle"])
        if not isinstance(mainDict[str(i+1)]["Length1"], float):
            print("Length1 is not correct for " + mainDict[str(i+1)]["Tag"])
            print(mainDict[str(i+1)]["Length1"])
        if not isinstance(mainDict[str(i+1)]["Length2"], float):
            print("Length2 is not correct for " + mainDict[str(i+1)]["Tag"])
            print(mainDict[str(i+1)]["Length2"])
    if not isinstance(mainDict["patientId"], int):
        print("patientId is not correct")
        print(mainDict["patientId"])
    print("Done")

    ##Write the Excel
    #Create the Excel and the sheet, and write the headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws['A1'] = "Tag"
    ws['A2'] = "Gantry Angle"
    ws['A3'] = "Length1(cm) UL"
    ws['A4'] = "Length2(cm) UL"
    ws['A5'] = "MU"
    ws['A6'] = "Y/X1/X2 (mm)"
    for row in range(mainDict["nbStrips"]):
        _ = ws.cell(column=1, row=row+7, value=str(row+1))

    #Write for all entries in mainDict
    for i in range(len(mainDict)-3):
        _ = ws.cell(column=3*i+2, row=1, value=mainDict[str(i+1)]["Sequence"])
        _ = ws.cell(column=3*i+2, row=2, value=mainDict[str(i+1)]["Tag"])
        _ = ws.cell(column=3*i+2, row=3, value=mainDict[str(i+1)]["Angle"])
        _ = ws.cell(column=3*i+2, row=4, value=mainDict[str(i+1)]["Length1"])
        _ = ws.cell(column=3*i+2, row=5, value=mainDict[str(i+1)]["Length2"])
        _ = ws.cell(column=3*i+2, row=6, value=mainDict[str(i+1)]["MU"])
        _ = ws.cell(column=3*i+2, row=7, value="Y")
        for row in range(mainDict["nbStrips"]):
            _ = ws.cell(column=3*i+2, row=row+8, value=str(mainDict[str(i+1)]["Y"][row]))
        _ = ws.cell(column=3*i+3, row=6, value="X1")
        for row in range(mainDict["nbStrips"]):
            _ = ws.cell(column=3*i+3, row=row+8, value=str(mainDict[str(i+1)]["X1"][row]))
        _ = ws.cell(column=3*i+4, row=6, value="X2")
        for row in range(mainDict["nbStrips"]):
            _ = ws.cell(column=3*i+4, row=row+8, value=str(mainDict[str(i+1)]["X2"][row]))

    #Save the Excel
    wb.save(filename = os.path.join(os.path.dirname(file), str(mainDict["patientId"]) + "_values.xlsx"))

    return(mainDict)

def convertToCorrectDict(mainDict, dataSet, fileNumber):
    # Set variables
    gantSpeed = 0
    doseRate = 1
    beam = 2
    seg = 3
    x1Diaphragm = 4
    x2Diaphragm = 5
    y1Diaphragm = 6
    y2Diaphragm = 7
    leaves = 8
    area = 9
    meanArea = 10
    angles = 11
    LSV = 12
    AAV = 13
    MCS = 14

    MUbeam = 0.0
    #start position
    tmpDictIndex = 0
    while mainDict[str(tmpDictIndex + 1)]["Sequence"] != fileNumber + 1:
        tmpDictIndex += 1
    dataSet[fileNumber][gantSpeed]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][gantSpeed]['Y'].append(0)
    dataSet[fileNumber][doseRate]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][doseRate]['Y'].append(mainDict[str(tmpDictIndex + 1)]["MU"])
    MUbeam += float(mainDict[str(tmpDictIndex + 1)]["MU"])
    dataSet[fileNumber][beam]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][beam]['Y'].append(0)
    dataSet[fileNumber][seg]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][seg]['Y'].append(0)
    dataSet[fileNumber][x1Diaphragm]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][x1Diaphragm]['Y'].append(mainDict[str(tmpDictIndex + 1)]["Length1"])
    dataSet[fileNumber][x2Diaphragm]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][x2Diaphragm]['Y'].append(mainDict[str(tmpDictIndex + 1)]["Length2"]*-1.0)
    dataSet[fileNumber][y1Diaphragm]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][y1Diaphragm]['Y'].append(2000)
    dataSet[fileNumber][y2Diaphragm]['X'].append(mainDict[str(tmpDictIndex + 1)]["Angle"])
    dataSet[fileNumber][y2Diaphragm]['Y'].append(2000)
    for row in range(mainDict["nbStrips"]):
        dataSet[fileNumber][leaves][0][row]['X'].append(mainDict[str(tmpDictIndex + 1)]["Y"][row])
        dataSet[fileNumber][leaves][0][row]['Y'].append(mainDict[str(tmpDictIndex + 1)]["X1"][row]*-100.0)
        dataSet[fileNumber][leaves][1][row]['X'].append(mainDict[str(tmpDictIndex + 1)]["Y"][row])
        dataSet[fileNumber][leaves][1][row]['Y'].append(mainDict[str(tmpDictIndex + 1)]["X2"][row]*-100.0)


    #all ended position
    for i in range(len(mainDict)-3):
        if "End" in mainDict[str(i+1)]["Tag"] and (fileNumber + 1 == mainDict[str(i+1)]["Sequence"]):
                dataSet[fileNumber][gantSpeed]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][gantSpeed]['Y'].append(0)
                dataSet[fileNumber][doseRate]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][doseRate]['Y'].append(mainDict[str(i+1)]["MU"])
                MUbeam += float(mainDict[str(i+1)]["MU"])
                dataSet[fileNumber][beam]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][beam]['Y'].append(0)
                dataSet[fileNumber][seg]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][seg]['Y'].append(0)
                dataSet[fileNumber][x1Diaphragm]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][x1Diaphragm]['Y'].append(mainDict[str(i+1)]["Length1"])
                dataSet[fileNumber][x2Diaphragm]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][x2Diaphragm]['Y'].append(mainDict[str(i+1)]["Length2"]*-1.0)
                dataSet[fileNumber][y1Diaphragm]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][y1Diaphragm]['Y'].append(2000)
                dataSet[fileNumber][y2Diaphragm]['X'].append(mainDict[str(i+1)]["Angle"])
                dataSet[fileNumber][y2Diaphragm]['Y'].append(2000)
                for row in range(mainDict["nbStrips"]):
                    dataSet[fileNumber][leaves][0][row]['X'].append(mainDict[str(i+1)]["Y"][row])
                    dataSet[fileNumber][leaves][0][row]['Y'].append(mainDict[str(i+1)]["X1"][row]*-100.0)
                    dataSet[fileNumber][leaves][1][row]['X'].append(mainDict[str(i+1)]["Y"][row])
                    dataSet[fileNumber][leaves][1][row]['Y'].append(mainDict[str(i+1)]["X2"][row]*-100.0)

    return((dataSet, MUbeam))
